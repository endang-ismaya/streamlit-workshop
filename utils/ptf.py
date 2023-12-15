import csv
import os
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font

redFill = PatternFill(
    start_color="FFFF0000",
    end_color="FFFF0000",
    fill_type="solid",
)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class PrintToFile:
    """Creating files with different types"""

    @staticmethod
    def to_csv(list_of_mo: list, file_to_save: str, csv_header: list) -> bool:
        """Create file to csv"""
        with open(file_to_save, "w", encoding="UTF8", newline="") as f:
            writer = csv.writer(f)

            # write the header
            writer.writerow(csv_header)

            # write the data
            for item in list_of_mo:
                writer.writerow(item)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_txt(contents: str, file_to_save: str) -> bool:
        """Create file to txt"""
        with open(file_to_save, "w") as f:
            f.write(contents)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_txt_unix(contents: str, file_to_save: str) -> bool:
        """Create file to txt unix"""
        with open(file_to_save, "w", newline="\n") as f:
            f.write(contents)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_txt_unix_append(contents: str, file_to_save: str) -> bool:
        """Create file to txt unix but append the file"""
        with open(file_to_save, "a", newline="\n") as f:
            f.write(contents)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_xlsx(
        file_to_save: str,
        ws_name: str,
        list_of_header: list,
        list_of_contents: list,
    ):
        """Create file to xlsx"""

        # check if workbook is exists
        if os.path.exists(file_to_save):
            wb = load_workbook(file_to_save)
        else:
            wb = Workbook()
            wb.remove(wb["Sheet"])

        # create cell
        ws_target = wb.create_sheet(ws_name)

        # starting row
        starting_row = 1

        # header
        for index, header in enumerate(list_of_header):
            ws_target.cell(row=starting_row, column=index + 1).value = header

        starting_row += 1
        for items in list_of_contents:
            for col, item in enumerate(items):
                ws_target.cell(row=starting_row, column=col + 1).value = item

            starting_row += 1

        wb.save(file_to_save)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_xlsx_undefined_filled(
        file_to_save: str,
        ws_name: str,
        list_of_header: list,
        list_of_contents: list,
        list_filled=["UNDEFINED", "MISMATCH", "None", "NONE"],
    ):
        """Create file to xlsx"""

        # check if workbook is exists
        if os.path.exists(file_to_save):
            wb = load_workbook(file_to_save)
        else:
            wb = Workbook()
            wb.remove(wb["Sheet"])

        # create cell
        ws_target = wb.create_sheet(ws_name)

        # starting row
        starting_row = 1

        # header
        for index, header in enumerate(list_of_header):
            ws_target.cell(row=starting_row, column=index + 1).value = header

        starting_row += 1
        for items in list_of_contents:
            for col, item in enumerate(items):
                ws_target.cell(row=starting_row, column=col + 1).value = item
                if item in list_filled:
                    ws_target.cell(row=starting_row, column=col + 1).fill = redFill

            starting_row += 1

        wb.save(file_to_save)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_xlsx_filled_and_col_offside(
        file_to_save: str,
        ws_name: str,
        list_of_header: list,
        list_of_contents: list,
        list_of_red: list = [],
        col_offside: int = 0,
    ):
        try:
            # check if workbook is exists
            if os.path.exists(file_to_save):
                wb = load_workbook(file_to_save)
            else:
                wb = Workbook()
                wb.remove(wb["Sheet"])

            # create cell
            if ws_name in wb.sheetnames:
                ws_target = wb[ws_name]
            else:
                ws_target = wb.create_sheet(ws_name)

            # starting row
            starting_row = 1

            # header
            for index, header in enumerate(list_of_header):
                col = index + 1 + col_offside
                ws_target.cell(row=starting_row, column=col).value = header

            starting_row += 1
            for items in list_of_contents:
                for col, item in enumerate(items):
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).value = item
                    if item in list_of_red:
                        ws_target.cell(
                            row=starting_row, column=col + 1 + col_offside
                        ).fill = redFill

                starting_row += 1

            wb.save(file_to_save)

            return True

        except Exception:
            errors = traceback.format_exc()
            print(errors)
            return False

    @staticmethod
    def to_xlsx_offside_noheader(
        file_to_save: str,
        ws_name: str,
        list_of_contents: list,
        list_of_red: list = [],
        col_offside: int = 0,
    ):
        try:
            # check if workbook is exists
            if os.path.exists(file_to_save):
                wb = load_workbook(file_to_save)
            else:
                wb = Workbook()
                wb.remove(wb["Sheet"])

            # create cell
            if ws_name in wb.sheetnames:
                ws_target = wb[ws_name]
            else:
                ws_target = wb.create_sheet(ws_name)

            # starting row
            starting_row = 2

            # header
            # for index, header in enumerate(list_of_header):
            #     col = index + 1 + col_offside
            #     ws_target.cell(row=starting_row, column=col).value = header

            starting_row += 1
            for items in list_of_contents:
                for col, item in enumerate(items):
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).value = item
                    if item in list_of_red:
                        ws_target.cell(
                            row=starting_row, column=col + 1 + col_offside
                        ).fill = redFill

                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).border = thin_border

                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).alignment = Alignment(horizontal="center", vertical="center")

                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).font = Font(name="Calibri", size=12)

                starting_row += 1

            wb.save(file_to_save)

            return True

        except Exception:
            errors = traceback.format_exc()
            print(errors)
            return False
